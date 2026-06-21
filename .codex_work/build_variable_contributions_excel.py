from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(
    "/Users/juniorjumbong/Desktop/personal-website/00_tds/01_PD_credit_scoring"
)
sys.path.insert(0, str(PROJECT_ROOT))

from src.modeling.score_computation import (
    build_score_outputs,
    load_score_model_artifact,
)


PICKLE_PATH = PROJECT_ROOT / "outputs" / "model_selection" / "best_4_variables_logit_scorecard.pkl"
TRAIN_PATH = PROJECT_ROOT / "data" / "train_discretized.csv"
TEST_PATH = PROJECT_ROOT / "data" / "test_discretized.csv"
OOT_PATH = PROJECT_ROOT / "data" / "oot_discretized.csv"
OUTPUT_PATH = PROJECT_ROOT / "outputs" / "model_selection" / "variable_score_contributions.xlsx"


RED = "9B1713"
DARK_RED = "7C100D"
LIGHT_GRAY = "EDEDED"
MID_GRAY = "C8C8C8"
WHITE = "FFFFFF"
BLACK = "222222"


def build_contribution_detail(train_df, modality_score_table, contributions, artifact):
    contribution_rows = []

    for variable_number, variable in enumerate(artifact["variables"], start=1):
        variable_scores = modality_score_table.loc[
            modality_score_table["variable"].eq(variable)
        ].sort_values("coefficient")
        contribution = contributions.loc[
            contributions["variable"].eq(variable)
        ].iloc[0]
        distribution = (
            train_df[variable]
            .astype(str)
            .value_counts(normalize=True, dropna=False)
            .to_dict()
        )
        first_row = True

        for _, score_row in variable_scores.iterrows():
            contribution_rows.append({
                "#": variable_number,
                "Variable": score_row["variable_label"],
                "Modality": score_row["modality_label"],
                "Repartition": distribution.get(str(score_row["modality"]), 0.0),
                "Score": score_row["score_points"],
                "Mean Score": contribution["weighted_mean_score"] if first_row else "",
                "Contribution Score": contribution["score_contribution_pct"] / 100 if first_row else "",
            })
            first_row = False

    return pd.DataFrame(contribution_rows)


def style_cell(cell, fill=None, font_color=BLACK, bold=False, size=11):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=font_color, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=WHITE),
        right=Side(style="thin", color=WHITE),
        top=Side(style="thin", color=WHITE),
        bottom=Side(style="thin", color=WHITE),
    )


def write_contribution_workbook(detail_df, contributions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Variable contributions"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Contribution des variables au score"
    style_cell(ws["A1"], fill=WHITE, font_color=RED, bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["#", "Variable", "Modality", "Repartition", "Score", "mean Score", "Contribution Score"]
    ws.append([])
    ws.append(headers)

    for cell in ws[3]:
        style_cell(cell, fill=RED, font_color=WHITE, bold=True, size=11)

    start_row = 4

    for row in detail_df.itertuples(index=False):
        ws.append(list(row))

    end_row = start_row + len(detail_df) - 1

    for row_idx in range(start_row, end_row + 1):
        row_fill = LIGHT_GRAY if row_idx % 2 == 0 else MID_GRAY

        for col_idx in range(1, 8):
            style_cell(ws.cell(row_idx, col_idx), fill=row_fill)

        style_cell(ws.cell(row_idx, 1), fill=DARK_RED, font_color=WHITE, bold=True, size=12)
        ws.cell(row_idx, 4).number_format = "0.00%"
        ws.cell(row_idx, 5).number_format = "0.00"
        ws.cell(row_idx, 6).number_format = "0.00"
        ws.cell(row_idx, 7).number_format = "0.00%"

    group_start = start_row

    while group_start <= end_row:
        variable = ws.cell(group_start, 2).value
        group_end = group_start

        while group_end + 1 <= end_row and ws.cell(group_end + 1, 2).value == variable:
            group_end += 1

        if group_end > group_start:
            ws.merge_cells(start_row=group_start, start_column=1, end_row=group_end, end_column=1)
            ws.merge_cells(start_row=group_start, start_column=2, end_row=group_end, end_column=2)
            ws.merge_cells(start_row=group_start, start_column=6, end_row=group_end, end_column=6)
            ws.merge_cells(start_row=group_start, start_column=7, end_row=group_end, end_column=7)

            for col_idx in [1, 2, 6, 7]:
                ws.cell(group_start, col_idx).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        group_start = group_end + 1

    widths = {
        "A": 8,
        "B": 28,
        "C": 30,
        "D": 18,
        "E": 12,
        "F": 16,
        "G": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 28

    for row_idx in range(start_row, end_row + 1):
        ws.row_dimensions[row_idx].height = 32

    ws.freeze_panes = "A4"

    summary = wb.create_sheet("Summary")
    summary.sheet_view.showGridLines = False
    summary.append(["Variable", "Max score", "Scale contribution", "Mean score", "Weighted SD", "Score contribution"])

    for cell in summary[1]:
        style_cell(cell, fill=RED, font_color=WHITE, bold=True, size=11)

    for row in contributions[
        [
            "variable_label",
            "max_score_points",
            "scale_contribution_pct",
            "weighted_mean_score",
            "weighted_sd_score",
            "score_contribution_pct",
        ]
    ].itertuples(index=False):
        summary.append(list(row))

    for row_idx in range(2, summary.max_row + 1):
        fill = LIGHT_GRAY if row_idx % 2 == 0 else MID_GRAY

        for col_idx in range(1, 7):
            style_cell(summary.cell(row_idx, col_idx), fill=fill)

        summary.cell(row_idx, 2).number_format = "0.00"
        summary.cell(row_idx, 3).number_format = "0.00"
        summary.cell(row_idx, 4).number_format = "0.00"
        summary.cell(row_idx, 5).number_format = "0.00"
        summary.cell(row_idx, 6).number_format = "0.00"

    for col_idx in range(1, 7):
        summary.column_dimensions[get_column_letter(col_idx)].width = 22

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


artifact = load_score_model_artifact(PICKLE_PATH)
train_df = pd.read_csv(TRAIN_PATH)
test_df = pd.read_csv(TEST_PATH)
oot_df = pd.read_csv(OOT_PATH)

modality_scores, _, variable_contributions = build_score_outputs(
    artifact=artifact,
    train_df=train_df,
    test_df=test_df,
    oot_df=oot_df,
)
detail = build_contribution_detail(
    train_df=train_df,
    modality_score_table=modality_scores,
    contributions=variable_contributions,
    artifact=artifact,
)
write_contribution_workbook(detail, variable_contributions)

print(OUTPUT_PATH)
